import shutil
import openpyxl

from src.currency_converter import rub_to_byn

CHAR_NAME = "Название_Характеристики"
CHAR_VALUE = "Измерение_Характеристики"
CHAR = "Значение_Характеристики"

BOATS_SEARCH = "Лодка, лодки, лодки breeze, BREEZE"
ENGINES_SEARCH = \
    "Лодочные моторы, моторы, двигатели," \
    "двигатели breeze, моторы breeze, BREEZE"

CATEGORY_BOATS = "https://deal.by/Lodki-"
CATEGORY_ENGINES = "https://deal.by/Lodochnye-motory"

GROUP_ENGINES = [10558128, "Лодочные моторы BREEZE"]
GROUP_BOATS = [10438662, "Лодки ПВХ"]


def copy_example_table():
    PATH = "src/data/import_example.xlsx"
    shutil.copy2(PATH, "import.xlsx")


def insert_data(data: dict, good_row: int):
    workbook = openpyxl.load_workbook("import.xlsx")
    sheet = workbook.active

    # Article
    sheet.cell(row=good_row, column=1, value=data["article"])
    sheet.cell(row=good_row, column=14, value=data["article"])
    sheet.cell(row=good_row, column=15, value=data["article"])

    # Good name & Description
    sheet.cell(row=good_row, column=2, value=data["good_name"])
    sheet.cell(row=good_row, column=4, value=data["description"])

    if data["category"] == "Моторы":
        sheet.cell(row=good_row, column=3, value=ENGINES_SEARCH)

        sheet.cell(row=good_row, column=11, value=GROUP_ENGINES[0])
        sheet.cell(row=good_row, column=12, value=GROUP_ENGINES[1])
        sheet.cell(row=good_row, column=13, value=CATEGORY_ENGINES)
    elif data["category"] == "Лодки":
        sheet.cell(row=good_row, column=3, value=BOATS_SEARCH)

        sheet.cell(row=good_row, column=11, value=GROUP_BOATS[0])
        sheet.cell(row=good_row, column=12, value=GROUP_BOATS[1])
        sheet.cell(row=good_row, column=13, value=CATEGORY_BOATS)
    else:
        print("ОШИБКА КАТЕГОРГИИ")

    if data["in_stock"] is True:
        sheet.cell(row=good_row, column=9, value="+")
    else:
        sheet.cell(row=good_row, column=9, value="-")

    # Price and unit
    sheet.cell(row=good_row, column=5, value=rub_to_byn(data["price"]))
    sheet.cell(row=good_row, column=6, value="BYN")
    sheet.cell(row=good_row, column=7, value="шт.")

    sheet.cell(row=good_row, column=8, value=data["image"])

    sheet.cell(row=good_row, column=10, value="BREEZE")

    # Chars
    max_chars = 16
    for k, v in data["chars"].items():
        sheet.cell(row=1, column=max_chars, value=CHAR_NAME)
        sheet.cell(row=1, column=max_chars + 1, value=CHAR_VALUE)
        sheet.cell(row=1, column=max_chars + 2, value=CHAR)

        sheet.cell(row=good_row, column=max_chars, value=k)
        sheet.cell(row=good_row, column=max_chars + 2, value=v)

        max_chars += 4

    workbook.save("import.xlsx")
    workbook.close()
